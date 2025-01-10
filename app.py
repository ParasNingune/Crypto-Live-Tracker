# Task-1 Fetch Data
import requests

# Fetching data from CoinGecko
def fetch_crypto_data():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 50,
        "page": 1,
        "sparkline": False
    }

    response = requests.get(url, params=params)

    if response.status_code == 200:
        data = response.json()
        return data
    else:
        print(f"Error: Unable to fetch data (Status Code: {response.status_code})")
        return None

# Print initial Details
crypto_data = fetch_crypto_data()
if crypto_data:
    print("Top 50 Cryptocurrencies:")
    for i, coin in enumerate(crypto_data, start=1):
        print(f"{i}. {coin['name']} ({coin['symbol'].upper()}) - ${coin['current_price']} {coin['market_cap']} {coin['total_volume']} {coin['price_change_percentage_24h']}")


# Task-2 Data Analysis
def analyze_crypto_data(data):
    
    # Identify the top 5 cryptocurrencies by market capitalization
    top_5_by_market_cap = sorted(data, key=lambda x: x['market_cap'], reverse=True)[:5]

    # Calculate the average price of the top 50 cryptocurrencies
    average_price = sum(coin['current_price'] for coin in data) / len(data)

    # Find the highest and lowest percentage change in 24 hours
    highest_change = max(data, key=lambda x: x['price_change_percentage_24h'])
    lowest_change = min(data, key=lambda x: x['price_change_percentage_24h'])

    # Display the analysis results
    print("\n--- Data Analysis ---")
    print("\nTop 5 Cryptocurrencies by Market Cap:")
    i=1
    for coin in top_5_by_market_cap:
        print(f"{i}. {coin['name']} ({coin['symbol'].upper()}) - Market Cap: ${coin['market_cap']:,}")
        i += 1

    print(f"\nAverage Price of Top 50 Cryptocurrencies: ${average_price:.2f}")
    print(f"\nHighest 24h % Change: {highest_change['name']} ({highest_change['price_change_percentage_24h']:.2f}%)")
    print(f"Lowest 24h % Change: {lowest_change['name']} ({lowest_change['price_change_percentage_24h']:.2f}%)")

analyze_crypto_data(crypto_data)


# Task-3 Live-Running Excel Sheet
import openpyxl
import time
from openpyxl.styles import PatternFill
import os
from fpdf import FPDF

def update_google_sheet(data):
   
    headers = ['Cryptocurrency Name', 'Symbol', 'Current Price (USD)', 'Market Capitalization',
               '24-hour Trading Volume', 'Price Change (24h, %)']
    
    existing_data = sheet.get_all_values()

    if len(existing_data) == 0:
        sheet.append_row(headers)
        existing_data = [headers]

    for coin in data:
        updated = False
        for i, row in enumerate(existing_data[1:], start=1):
            if row[1].upper() == coin['symbol'].upper():
                
                existing_data[i] = [
                    coin['name'],
                    coin['symbol'].upper(),
                    coin['current_price'],
                    coin['market_cap'],
                    coin['total_volume'],
                    coin['price_change_percentage_24h']
                ]
                updated = True
                break

        if not updated:
            existing_data.append([
                coin['name'],
                coin['symbol'].upper(),
                coin['current_price'],
                coin['market_cap'],
                coin['total_volume'],
                coin['price_change_percentage_24h']
            ])

    sheet.clear()
    sheet.append_rows(existing_data, value_input_option="USER_ENTERED")


def update_excel_sheet(data):
    file_path = 'crypto_data.xlsx'
    
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    else:
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        
        sheet['A1'] = 'Cryptocurrency Name'
        sheet['B1'] = 'Symbol'
        sheet['C1'] = 'Current Price (USD)'
        sheet['D1'] = 'Market Capitalization'
        sheet['E1'] = '24-hour Trading Volume'
        sheet['F1'] = 'Price Change (24h, %)'

    # Create fill styles for row highlighting
    red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')  # Red for decrease
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Green for increase

    # Write data for each cryptocurrency
    row = sheet.max_row + 1
    for coin in data:
        sheet[f'A{row}'] = coin['name']
        sheet[f'B{row}'] = coin['symbol'].upper()
        sheet[f'C{row}'] = coin['current_price']
        sheet[f'D{row}'] = coin['market_cap']
        sheet[f'E{row}'] = coin['total_volume']
        sheet[f'F{row}'] = coin['price_change_percentage_24h']


        price_change = coin['price_change_percentage_24h']
        
        if price_change < 0:
            for col in range(1, 7):
                sheet.cell(row=row, column=col).fill = red_fill
        else:
            
            for col in range(1, 7):  # Columns A to F
                sheet.cell(row=row, column=col).fill = green_fill

        row += 1
    
    workbook.save(file_path)

def generate_analysis_report(data):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font('Arial', 'B', 16)
    pdf.cell(200, 10, txt="Cryptocurrency Market Analysis", ln=True, align='C')

    pdf.set_font('Arial', '', 12)
    pdf.ln(10)
    pdf.cell(200, 10, txt="Key Insights and Analysis", ln=True)

    pdf.ln(10)
    pdf.set_font('Arial', '', 10)
    pdf.cell(30, 10, 'Name', border=1)
    pdf.cell(30, 10, 'Symbol', border=1)
    pdf.cell(50, 10, 'Current Price (USD)', border=1)
    pdf.cell(50, 10, '24h Change (%)', border=1)
    pdf.ln(10)

    for coin in data:
        pdf.cell(30, 10, coin['name'], border=1)
        pdf.cell(30, 10, coin['symbol'], border=1)
        pdf.cell(50, 10, f"${coin['current_price']}", border=1)
        pdf.cell(50, 10, f"{coin['price_change_percentage_24h']}%", border=1)
        pdf.ln(10)

    pdf.output("analysis_report.pdf")


def run_live_updates():
    while True:
        crypto_data = fetch_crypto_data()
        if crypto_data:
            update_excel_sheet(crypto_data)
            generate_analysis_report(crypto_data)
        time.sleep(300)  # Wait for 5 minutes

run_live_updates()